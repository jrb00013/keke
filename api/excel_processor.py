import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import json
import csv
import io
import logging
from typing import Dict, List, Optional, Tuple, Any, Union
from datetime import datetime
import re
from pathlib import Path

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelProcessor:
    """
    Advanced Excel file processor with data analysis, manipulation, and enhancement capabilities
    """
    
    def __init__(self):
        self.workbook = None
        self.dataframes = {}
        self.current_sheet = None
        self.supported_formats = ['.xlsx', '.xls', '.csv', '.json']
        
    def load_file(self, file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """
        Load Excel file and return metadata
        """
        try:
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext == '.csv':
                return self._load_csv(file_path)
            elif file_ext == '.json':
                return self._load_json(file_path)
            elif file_ext in ['.xlsx', '.xls']:
                return self._load_excel(file_path, sheet_name)
            else:
                raise ValueError(f"Unsupported file format: {file_ext}")
                
        except Exception as e:
            logger.error(f"Error loading file {file_path}: {str(e)}")
            raise
    
    def _load_excel(self, file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """Load Excel file"""
        self.workbook = load_workbook(file_path, data_only=True)
        
        # Get all sheet names
        sheet_names = self.workbook.sheetnames
        
        # Load dataframes for each sheet
        excel_data = pd.read_excel(file_path, sheet_name=None)
        
        result = {
            'file_path': file_path,
            'sheet_names': sheet_names,
            'sheets': {},
            'total_sheets': len(sheet_names),
            'loaded_at': datetime.now().isoformat()
        }
        
        for name, df in excel_data.items():
            self.dataframes[name] = df
            result['sheets'][name] = {
                'rows': len(df),
                'columns': len(df.columns),
                'column_names': df.columns.tolist(),
                'data_types': df.dtypes.to_dict(),
                'has_nulls': df.isnull().sum().to_dict(),
                'memory_usage': df.memory_usage(deep=True).sum()
            }
        
        return result
    
    def _load_csv(self, file_path: str) -> Dict[str, Any]:
        """Load CSV file"""
        df = pd.read_csv(file_path)
        self.dataframes['Sheet1'] = df
        
        return {
            'file_path': file_path,
            'sheet_names': ['Sheet1'],
            'sheets': {
                'Sheet1': {
                    'rows': len(df),
                    'columns': len(df.columns),
                    'column_names': df.columns.tolist(),
                    'data_types': df.dtypes.to_dict(),
                    'has_nulls': df.isnull().sum().to_dict(),
                    'memory_usage': df.memory_usage(deep=True).sum()
                }
            },
            'total_sheets': 1,
            'loaded_at': datetime.now().isoformat()
        }
    
    def _load_json(self, file_path: str) -> Dict[str, Any]:
        """Load JSON file"""
        with open(file_path, 'r') as f:
            data = json.load(f)
        
        if isinstance(data, list):
            df = pd.DataFrame(data)
        elif isinstance(data, dict):
            df = pd.DataFrame([data])
        else:
            raise ValueError("JSON must be a list of objects or a single object")
        
        self.dataframes['Sheet1'] = df
        
        return {
            'file_path': file_path,
            'sheet_names': ['Sheet1'],
            'sheets': {
                'Sheet1': {
                    'rows': len(df),
                    'columns': len(df.columns),
                    'column_names': df.columns.tolist(),
                    'data_types': df.dtypes.to_dict(),
                    'has_nulls': df.isnull().sum().to_dict(),
                    'memory_usage': df.memory_usage(deep=True).sum()
                }
            },
            'total_sheets': 1,
            'loaded_at': datetime.now().isoformat()
        }
    
    def analyze_data(self, sheet_name: str) -> Dict[str, Any]:
        """
        Perform comprehensive data analysis on a sheet
        """
        if sheet_name not in self.dataframes:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        
        df = self.dataframes[sheet_name]
        
        analysis = {
            'basic_info': {
                'rows': len(df),
                'columns': len(df.columns),
                'memory_usage': df.memory_usage(deep=True).sum(),
                'shape': df.shape
            },
            'data_quality': {
                'null_counts': df.isnull().sum().to_dict(),
                'duplicate_rows': df.duplicated().sum(),
                'empty_rows': df.isnull().all(axis=1).sum(),
                'data_types': df.dtypes.to_dict()
            },
            'statistics': {},
            'patterns': {},
            'recommendations': []
        }
        
        # Statistical analysis for numeric columns
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) > 0:
            analysis['statistics']['numeric'] = df[numeric_cols].describe().to_dict()
        
        # Categorical analysis
        categorical_cols = df.select_dtypes(include=['object']).columns
        if len(categorical_cols) > 0:
            analysis['statistics']['categorical'] = {}
            for col in categorical_cols:
                analysis['statistics']['categorical'][col] = {
                    'unique_values': df[col].nunique(),
                    'most_common': df[col].value_counts().head(5).to_dict(),
                    'null_percentage': (df[col].isnull().sum() / len(df)) * 100
                }
        
        # Pattern detection
        analysis['patterns'] = self._detect_patterns(df)
        
        # Generate recommendations
        analysis['recommendations'] = self._generate_recommendations(df, analysis)
        
        return analysis
    
    def _detect_patterns(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Detect patterns in the data"""
        patterns = {
            'potential_duplicates': [],
            'outliers': {},
            'trends': {},
            'correlations': {}
        }
        
        # Detect potential duplicates
        duplicate_mask = df.duplicated(keep=False)
        if duplicate_mask.any():
            patterns['potential_duplicates'] = df[duplicate_mask].index.tolist()
        
        # Detect outliers in numeric columns
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        for col in numeric_cols:
            Q1 = df[col].quantile(0.25)
            Q3 = df[col].quantile(0.75)
            IQR = Q3 - Q1
            lower_bound = Q1 - 1.5 * IQR
            upper_bound = Q3 + 1.5 * IQR
            
            outliers = df[(df[col] < lower_bound) | (df[col] > upper_bound)]
            if len(outliers) > 0:
                patterns['outliers'][col] = {
                    'count': len(outliers),
                    'indices': outliers.index.tolist(),
                    'values': outliers[col].tolist()
                }
        
        # Calculate correlations
        if len(numeric_cols) > 1:
            patterns['correlations'] = df[numeric_cols].corr().to_dict()
        
        return patterns
    
    def _generate_recommendations(self, df: pd.DataFrame, analysis: Dict) -> List[str]:
        """Generate data improvement recommendations"""
        recommendations = []
        
        # Check for null values
        null_counts = analysis['data_quality']['null_counts']
        high_null_cols = [col for col, count in null_counts.items() if count > len(df) * 0.1]
        if high_null_cols:
            recommendations.append(f"Consider handling null values in columns: {', '.join(high_null_cols)}")
        
        # Check for duplicates
        if analysis['data_quality']['duplicate_rows'] > 0:
            recommendations.append(f"Found {analysis['data_quality']['duplicate_rows']} duplicate rows. Consider removing duplicates.")
        
        # Check data types
        for col, dtype in analysis['data_quality']['data_types'].items():
            if dtype == 'object':
                # Check if it could be converted to numeric
                try:
                    pd.to_numeric(df[col], errors='raise')
                    recommendations.append(f"Column '{col}' appears to contain numeric data but is stored as text. Consider converting to numeric.")
                except:
                    pass
        
        # Check for outliers
        if analysis['patterns']['outliers']:
            outlier_cols = list(analysis['patterns']['outliers'].keys())
            recommendations.append(f"Outliers detected in columns: {', '.join(outlier_cols)}. Review for data quality issues.")
        
        return recommendations
    
    def clean_data(self, sheet_name: str, operations: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Clean data based on specified operations
        """
        if sheet_name not in self.dataframes:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        
        df = self.dataframes[sheet_name].copy()
        original_shape = df.shape
        
        results = {
            'original_shape': original_shape,
            'operations_applied': [],
            'final_shape': None,
            'changes_summary': {}
        }
        
        for operation in operations:
            op_type = operation.get('type')
            params = operation.get('params', {})
            
            if op_type == 'remove_duplicates':
                df = df.drop_duplicates()
                results['operations_applied'].append('remove_duplicates')
                
            elif op_type == 'remove_nulls':
                if params.get('strategy') == 'drop_rows':
                    df = df.dropna()
                elif params.get('strategy') == 'drop_columns':
                    threshold = params.get('threshold', 0.5)
                    df = df.dropna(thresh=int(len(df) * (1 - threshold)))
                elif params.get('strategy') == 'fill':
                    method = params.get('method', 'forward')
                    df = df.fillna(method=method)
                results['operations_applied'].append('remove_nulls')
                
            elif op_type == 'convert_types':
                for col, new_type in params.get('conversions', {}).items():
                    if col in df.columns:
                        try:
                            if new_type == 'numeric':
                                df[col] = pd.to_numeric(df[col], errors='coerce')
                            elif new_type == 'datetime':
                                df[col] = pd.to_datetime(df[col], errors='coerce')
                            elif new_type == 'string':
                                df[col] = df[col].astype(str)
                        except Exception as e:
                            logger.warning(f"Could not convert column {col} to {new_type}: {e}")
                results['operations_applied'].append('convert_types')
                
            elif op_type == 'rename_columns':
                df = df.rename(columns=params.get('mapping', {}))
                results['operations_applied'].append('rename_columns')
                
            elif op_type == 'filter_rows':
                condition = params.get('condition')
                if condition:
                    df = df.query(condition)
                results['operations_applied'].append('filter_rows')
        
        # Update the dataframe
        self.dataframes[sheet_name] = df
        results['final_shape'] = df.shape
        results['changes_summary'] = {
            'rows_removed': original_shape[0] - df.shape[0],
            'columns_removed': original_shape[1] - df.shape[1]
        }
        
        return results
    
    def create_chart(self, sheet_name: str, chart_config: Dict[str, Any]) -> str:
        """
        Create a chart and return the chart data
        """
        if sheet_name not in self.dataframes:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        
        df = self.dataframes[sheet_name]
        chart_type = chart_config.get('type', 'bar')
        
        # Create a new workbook for the chart
        wb = Workbook()
        ws = wb.active
        ws.title = f"{sheet_name}_chart"
        
        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Create chart based on type
        if chart_type == 'bar':
            chart = BarChart()
        elif chart_type == 'line':
            chart = LineChart()
        elif chart_type == 'pie':
            chart = PieChart()
        else:
            raise ValueError(f"Unsupported chart type: {chart_type}")
        
        # Configure chart
        chart.title = chart_config.get('title', f'{chart_type.title()} Chart')
        chart.style = chart_config.get('style', 10)
        
        # Add data to chart
        x_col = chart_config.get('x_column', df.columns[0])
        y_cols = chart_config.get('y_columns', [df.columns[1]] if len(df.columns) > 1 else [df.columns[0]])
        
        if isinstance(y_cols, str):
            y_cols = [y_cols]
        
        # Find column indices
        x_idx = df.columns.get_loc(x_col) + 1  # +1 for Excel 1-based indexing
        y_indices = [df.columns.get_loc(col) + 1 for col in y_cols]
        
        # Create data reference
        data = Reference(ws, min_col=min(y_indices), min_row=1, max_col=max(y_indices), max_row=len(df) + 1)
        categories = Reference(ws, min_col=x_idx, min_row=2, max_row=len(df) + 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        # Add chart to worksheet
        ws.add_chart(chart, "E2")
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def export_data(self, sheet_name: str, format: str, **kwargs) -> bytes:
        """
        Export data in various formats
        """
        if sheet_name not in self.dataframes:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        
        df = self.dataframes[sheet_name]
        output = io.BytesIO()
        
        if format.lower() == 'csv':
            df.to_csv(output, index=False, **kwargs)
        elif format.lower() == 'json':
            df.to_json(output, orient='records', **kwargs)
        elif format.lower() == 'excel':
            df.to_excel(output, index=False, **kwargs)
        elif format.lower() == 'parquet':
            df.to_parquet(output, **kwargs)
        else:
            raise ValueError(f"Unsupported export format: {format}")
        
        output.seek(0)
        return output.getvalue()
    
    def apply_formulas(self, sheet_name: str, formulas: Dict[str, str]) -> Dict[str, Any]:
        """
        Apply Excel-like formulas to the data
        """
        if sheet_name not in self.dataframes:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        
        df = self.dataframes[sheet_name].copy()
        results = {}
        
        for column, formula in formulas.items():
            try:
                # Parse and apply formula
                result = self._evaluate_formula(df, formula)
                df[column] = result
                results[column] = {
                    'success': True,
                    'formula': formula,
                    'result_type': type(result).__name__
                }
            except Exception as e:
                results[column] = {
                    'success': False,
                    'formula': formula,
                    'error': str(e)
                }
        
        # Update the dataframe
        self.dataframes[sheet_name] = df
        
        return results
    
    def _evaluate_formula(self, df: pd.DataFrame, formula: str) -> Any:
        """
        Evaluate Excel-like formulas
        """
        # Simple formula evaluation (can be extended)
        formula = formula.strip()
        
        if formula.startswith('='):
            formula = formula[1:]
        
        # Handle common Excel functions
        if formula.upper().startswith('SUM('):
            # Extract column reference
            match = re.search(r'SUM\(([A-Z]+):([A-Z]+)\)', formula.upper())
            if match:
                start_col, end_col = match.groups()
                start_idx = self._column_letter_to_index(start_col)
                end_idx = self._column_letter_to_index(end_col)
                return df.iloc[:, start_idx:end_idx+1].sum(axis=1)
        
        elif formula.upper().startswith('AVERAGE('):
            match = re.search(r'AVERAGE\(([A-Z]+):([A-Z]+)\)', formula.upper())
            if match:
                start_col, end_col = match.groups()
                start_idx = self._column_letter_to_index(start_col)
                end_idx = self._column_letter_to_index(end_col)
                return df.iloc[:, start_idx:end_idx+1].mean(axis=1)
        
        elif formula.upper().startswith('COUNT('):
            match = re.search(r'COUNT\(([A-Z]+):([A-Z]+)\)', formula.upper())
            if match:
                start_col, end_col = match.groups()
                start_idx = self._column_letter_to_index(start_col)
                end_idx = self._column_letter_to_index(end_col)
                return df.iloc[:, start_idx:end_idx+1].count(axis=1)
        
        # Handle column references like A1, B2, etc.
        elif re.match(r'^[A-Z]+[0-9]+$', formula.upper()):
            match = re.match(r'^([A-Z]+)([0-9]+)$', formula.upper())
            if match:
                col_letter, row_num = match.groups()
                col_idx = self._column_letter_to_index(col_letter)
                row_idx = int(row_num) - 1  # Convert to 0-based
                if col_idx < len(df.columns) and row_idx < len(df):
                    return df.iloc[row_idx, col_idx]
        
        # Handle arithmetic operations
        elif re.match(r'^[A-Z]+[0-9]+[\+\-\*\/][A-Z]+[0-9]+$', formula.upper()):
            # Simple arithmetic between cells
            parts = re.split(r'([\+\-\*\/])', formula.upper())
            if len(parts) == 3:
                cell1, operator, cell2 = parts
                val1 = self._get_cell_value(df, cell1)
                val2 = self._get_cell_value(df, cell2)
                
                if operator == '+':
                    return val1 + val2
                elif operator == '-':
                    return val1 - val2
                elif operator == '*':
                    return val1 * val2
                elif operator == '/':
                    return val1 / val2
        
        # If no pattern matches, try to evaluate as Python expression
        try:
            # Replace column references with actual values
            safe_formula = self._make_formula_safe(df, formula)
            return eval(safe_formula)
        except:
            raise ValueError(f"Unable to evaluate formula: {formula}")
    
    def _column_letter_to_index(self, letter: str) -> int:
        """Convert Excel column letter to index"""
        result = 0
        for char in letter:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result - 1
    
    def _get_cell_value(self, df: pd.DataFrame, cell_ref: str) -> Any:
        """Get value from cell reference like A1, B2, etc."""
        match = re.match(r'^([A-Z]+)([0-9]+)$', cell_ref.upper())
        if match:
            col_letter, row_num = match.groups()
            col_idx = self._column_letter_to_index(col_letter)
            row_idx = int(row_num) - 1
            if col_idx < len(df.columns) and row_idx < len(df):
                return df.iloc[row_idx, col_idx]
        return 0
    
    def _make_formula_safe(self, df: pd.DataFrame, formula: str) -> str:
        """Make formula safe for evaluation by replacing references"""
        # This is a simplified version - in production, you'd want more robust parsing
        safe_formula = formula
        for i, col in enumerate(df.columns):
            safe_formula = safe_formula.replace(f'COL_{i}', f'df.iloc[:, {i}]')
        return safe_formula
    
    def get_summary(self) -> Dict[str, Any]:
        """
        Get summary of all loaded data
        """
        summary = {
            'total_sheets': len(self.dataframes),
            'sheets': {},
            'total_rows': 0,
            'total_columns': 0,
            'total_memory_usage': 0
        }
        
        for name, df in self.dataframes.items():
            sheet_info = {
                'rows': len(df),
                'columns': len(df.columns),
                'memory_usage': df.memory_usage(deep=True).sum(),
                'data_types': df.dtypes.to_dict(),
                'null_counts': df.isnull().sum().to_dict()
            }
            summary['sheets'][name] = sheet_info
            summary['total_rows'] += len(df)
            summary['total_columns'] += len(df.columns)
            summary['total_memory_usage'] += df.memory_usage(deep=True).sum()
        
        return summary
    
    def get_preview(self, sheet_name: str, limit: int = 50) -> List[Dict[str, Any]]:
        """
        Get a preview of the data in a sheet
        """
        if sheet_name not in self.dataframes:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        
        df = self.dataframes[sheet_name]
        
        # Get first 'limit' rows
        preview_df = df.head(limit)
        
        # Convert to list of dictionaries
        return preview_df.to_dict('records')
    
    def get_columns(self, sheet_name: str) -> Dict[str, Any]:
        """
        Get detailed information about columns in a sheet
        """
        if sheet_name not in self.dataframes:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        
        df = self.dataframes[sheet_name]
        
        column_info = {}
        for col in df.columns:
            col_data = df[col]
            column_info[col] = {
                'data_type': str(col_data.dtype),
                'null_count': col_data.isnull().sum(),
                'null_percentage': (col_data.isnull().sum() / len(df)) * 100,
                'unique_count': col_data.nunique(),
                'unique_percentage': (col_data.nunique() / len(df)) * 100,
                'sample_values': col_data.dropna().head(5).tolist()
            }
            
            # Add statistical info for numeric columns
            if pd.api.types.is_numeric_dtype(col_data):
                column_info[col].update({
                    'min': col_data.min(),
                    'max': col_data.max(),
                    'mean': col_data.mean(),
                    'median': col_data.median(),
                    'std': col_data.std()
                })
        
        return column_info
    
    def validate_data(self, sheet_name: str, rules: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Validate data based on specified rules
        """
        if sheet_name not in self.dataframes:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        
        df = self.dataframes[sheet_name]
        validation_results = {
            'total_rows': len(df),
            'rules_applied': [],
            'violations': [],
            'summary': {
                'passed': 0,
                'failed': 0,
                'total_checks': 0
            }
        }
        
        for rule in rules:
            rule_type = rule.get('type')
            params = rule.get('params', {})
            
            if rule_type == 'not_null':
                columns = params.get('columns', [])
                for col in columns:
                    if col in df.columns:
                        null_count = df[col].isnull().sum()
                        validation_results['total_checks'] += 1
                        if null_count == 0:
                            validation_results['summary']['passed'] += 1
                        else:
                            validation_results['summary']['failed'] += 1
                            validation_results['violations'].append({
                                'rule': f'Column {col} should not be null',
                                'column': col,
                                'violation_count': null_count,
                                'violation_percentage': (null_count / len(df)) * 100
                            })
            
            elif rule_type == 'unique':
                columns = params.get('columns', [])
                for col in columns:
                    if col in df.columns:
                        duplicate_count = df[col].duplicated().sum()
                        validation_results['total_checks'] += 1
                        if duplicate_count == 0:
                            validation_results['summary']['passed'] += 1
                        else:
                            validation_results['summary']['failed'] += 1
                            validation_results['violations'].append({
                                'rule': f'Column {col} should be unique',
                                'column': col,
                                'violation_count': duplicate_count,
                                'violation_percentage': (duplicate_count / len(df)) * 100
                            })
            
            elif rule_type == 'range':
                column = params.get('column')
                min_val = params.get('min')
                max_val = params.get('max')
                
                if column in df.columns and pd.api.types.is_numeric_dtype(df[column]):
                    violations = df[(df[column] < min_val) | (df[column] > max_val)]
                    violation_count = len(violations)
                    validation_results['total_checks'] += 1
                    
                    if violation_count == 0:
                        validation_results['summary']['passed'] += 1
                    else:
                        validation_results['summary']['failed'] += 1
                        validation_results['violations'].append({
                            'rule': f'Column {column} should be between {min_val} and {max_val}',
                            'column': column,
                            'violation_count': violation_count,
                            'violation_percentage': (violation_count / len(df)) * 100
                        })
            
            elif rule_type == 'pattern':
                column = params.get('column')
                pattern = params.get('pattern')
                
                if column in df.columns:
                    # Convert to string for pattern matching
                    str_col = df[column].astype(str)
                    violations = ~str_col.str.match(pattern, na=False)
                    violation_count = violations.sum()
                    validation_results['total_checks'] += 1
                    
                    if violation_count == 0:
                        validation_results['summary']['passed'] += 1
                    else:
                        validation_results['summary']['failed'] += 1
                        validation_results['violations'].append({
                            'rule': f'Column {column} should match pattern {pattern}',
                            'column': column,
                            'violation_count': violation_count,
                            'violation_percentage': (violation_count / len(df)) * 100
                        })
            
            validation_results['rules_applied'].append(rule_type)
        
        return validation_results
    
    def transform_data(self, sheet_name: str, transformations: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Apply data transformations to a sheet
        """
        if sheet_name not in self.dataframes:
            raise ValueError(f"Sheet '{sheet_name}' not found")
        
        df = self.dataframes[sheet_name].copy()
        original_shape = df.shape
        
        results = {
            'original_shape': original_shape,
            'transformations_applied': [],
            'final_shape': None,
            'changes_summary': {}
        }
        
        for transformation in transformations:
            trans_type = transformation.get('type')
            params = transformation.get('params', {})
            
            if trans_type == 'pivot':
                index_cols = params.get('index', [])
                columns_col = params.get('columns')
                values_cols = params.get('values', [])
                agg_func = params.get('aggfunc', 'sum')
                
                if index_cols and columns_col and values_cols:
                    df = df.pivot_table(
                        index=index_cols,
                        columns=columns_col,
                        values=values_cols,
                        aggfunc=agg_func
                    )
                    results['transformations_applied'].append('pivot')
            
            elif trans_type == 'groupby':
                group_cols = params.get('groupby', [])
                agg_dict = params.get('aggregations', {})
                
                if group_cols and agg_dict:
                    df = df.groupby(group_cols).agg(agg_dict).reset_index()
                    results['transformations_applied'].append('groupby')
            
            elif trans_type == 'merge':
                other_sheet = params.get('other_sheet')
                merge_keys = params.get('keys', [])
                merge_type = params.get('type', 'inner')
                
                if other_sheet in self.dataframes and merge_keys:
                    other_df = self.dataframes[other_sheet]
                    df = df.merge(other_df, on=merge_keys, how=merge_type)
                    results['transformations_applied'].append('merge')
            
            elif trans_type == 'concat':
                other_sheets = params.get('other_sheets', [])
                axis = params.get('axis', 0)  # 0 for rows, 1 for columns
                
                other_dfs = [self.dataframes[sheet] for sheet in other_sheets if sheet in self.dataframes]
                if other_dfs:
                    df = pd.concat([df] + other_dfs, axis=axis)
                    results['transformations_applied'].append('concat')
            
            elif trans_type == 'sort':
                sort_cols = params.get('columns', [])
                ascending = params.get('ascending', True)
                
                if sort_cols:
                    df = df.sort_values(by=sort_cols, ascending=ascending)
                    results['transformations_applied'].append('sort')
            
            elif trans_type == 'filter':
                condition = params.get('condition')
                
                if condition:
                    df = df.query(condition)
                    results['transformations_applied'].append('filter')
        
        # Update the dataframe
        self.dataframes[sheet_name] = df
        results['final_shape'] = df.shape
        results['changes_summary'] = {
            'rows_changed': original_shape[0] - df.shape[0],
            'columns_changed': original_shape[1] - df.shape[1]
        }
        
        return results


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python excel_processor.py <command> [args...]")
        sys.exit(1)
    
    command = sys.argv[1]
    processor = ExcelProcessor()
    
    try:
        if command == "load_file":
            if len(sys.argv) < 3:
                print("Usage: python excel_processor.py load_file <file_path>")
                sys.exit(1)
            
            file_path = sys.argv[2]
            result = processor.load_file(file_path)
            print(json.dumps(result, indent=2))
            
        elif command == "analyze_data":
            if len(sys.argv) < 4:
                print("Usage: python excel_processor.py analyze_data <session_id> <sheet_name>")
                sys.exit(1)
            
            session_id = sys.argv[2]
            sheet_name = sys.argv[3]
            result = processor.analyze_data(sheet_name)
            print(json.dumps(result, indent=2))
            
        elif command == "clean_data":
            if len(sys.argv) < 5:
                print("Usage: python excel_processor.py clean_data <session_id> <sheet_name> <operations_json>")
                sys.exit(1)
            
            session_id = sys.argv[2]
            sheet_name = sys.argv[3]
            operations = json.loads(sys.argv[4])
            result = processor.clean_data(sheet_name, operations)
            print(json.dumps(result, indent=2))
            
        elif command == "create_chart":
            if len(sys.argv) < 5:
                print("Usage: python excel_processor.py create_chart <session_id> <sheet_name> <chart_config_json>")
                sys.exit(1)
            
            session_id = sys.argv[2]
            sheet_name = sys.argv[3]
            chart_config = json.loads(sys.argv[4])
            result = processor.create_chart(sheet_name, chart_config)
            print(result)
            
        elif command == "export_data":
            if len(sys.argv) < 5:
                print("Usage: python excel_processor.py export_data <session_id> <sheet_name> <format>")
                sys.exit(1)
            
            session_id = sys.argv[2]
            sheet_name = sys.argv[3]
            format_type = sys.argv[4]
            result = processor.export_data(sheet_name, format_type)
            print(result)
            
        elif command == "apply_formulas":
            if len(sys.argv) < 5:
                print("Usage: python excel_processor.py apply_formulas <session_id> <sheet_name> <formulas_json>")
                sys.exit(1)
            
            session_id = sys.argv[2]
            sheet_name = sys.argv[3]
            formulas = json.loads(sys.argv[4])
            result = processor.apply_formulas(sheet_name, formulas)
            print(json.dumps(result, indent=2))
            
        elif command == "get_summary":
            if len(sys.argv) < 3:
                print("Usage: python excel_processor.py get_summary <session_id>")
                sys.exit(1)
            
            session_id = sys.argv[2]
            result = processor.get_summary()
            print(json.dumps(result, indent=2))
            
        elif command == "get_preview":
            if len(sys.argv) < 5:
                print("Usage: python excel_processor.py get_preview <session_id> <sheet_name> <limit>")
                sys.exit(1)
            
            session_id = sys.argv[2]
            sheet_name = sys.argv[3]
            limit = int(sys.argv[4])
            result = processor.get_preview(sheet_name, limit)
            print(json.dumps(result, indent=2))
            
        elif command == "get_columns":
            if len(sys.argv) < 4:
                print("Usage: python excel_processor.py get_columns <session_id> <sheet_name>")
                sys.exit(1)
            
            session_id = sys.argv[2]
            sheet_name = sys.argv[3]
            result = processor.get_columns(sheet_name)
            print(json.dumps(result, indent=2))
            
        elif command == "validate_data":
            if len(sys.argv) < 5:
                print("Usage: python excel_processor.py validate_data <session_id> <sheet_name> <rules_json>")
                sys.exit(1)
            
            session_id = sys.argv[2]
            sheet_name = sys.argv[3]
            rules = json.loads(sys.argv[4])
            result = processor.validate_data(sheet_name, rules)
            print(json.dumps(result, indent=2))
            
        elif command == "transform_data":
            if len(sys.argv) < 5:
                print("Usage: python excel_processor.py transform_data <session_id> <sheet_name> <transformations_json>")
                sys.exit(1)
            
            session_id = sys.argv[2]
            sheet_name = sys.argv[3]
            transformations = json.loads(sys.argv[4])
            result = processor.transform_data(sheet_name, transformations)
            print(json.dumps(result, indent=2))
            
        else:
            print(f"Unknown command: {command}")
            sys.exit(1)
            
    except Exception as e:
        print(f"Error: {str(e)}", file=sys.stderr)
        sys.exit(1)